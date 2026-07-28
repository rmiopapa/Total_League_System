from pathlib import Path
from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, file_path):
        self.file_path = Path(file_path)

    def _read_xls_rows(self, max_rows=None):
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(
                "Excel 97-2003形式（.xls）の読み込みには xlrd が必要です。"
                "requirements.txt を更新してから再実行してください。"
            ) from exc

        workbook = xlrd.open_workbook(self.file_path)
        sheet = workbook.sheet_by_index(0)

        def convert(cell):
            if cell.ctype == xlrd.XL_CELL_DATE:
                return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
            if cell.ctype == xlrd.XL_CELL_BOOLEAN:
                return bool(cell.value)
            if cell.ctype == xlrd.XL_CELL_ERROR:
                return None
            if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
                return int(cell.value)
            return cell.value

        row_limit = sheet.nrows if max_rows is None else min(max_rows, sheet.nrows)
        rows = [
            [convert(sheet.cell(row_index, col_index)) for col_index in range(sheet.ncols)]
            for row_index in range(row_limit)
        ]
        return sheet, rows

    def read_summary(self):
        if self.file_path.suffix.lower() == ".xls":
            sheet, rows = self._read_xls_rows(max_rows=30)
            return {
                "file_name": self.file_path.name,
                "university_name": self.file_path.stem,
                "sheet_name": sheet.name,
                "max_row": sheet.nrows,
                "max_column": sheet.ncols,
                "rows": rows
            }

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        rows = []

        for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True):
            rows.append(list(row))

        summary = {
            "file_name": self.file_path.name,
            "university_name": self.file_path.stem,
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "rows": rows
        }

        wb.close()
        return summary

    def read_all_rows(self):
        if self.file_path.suffix.lower() == ".xls":
            _sheet, rows = self._read_xls_rows()
            return rows

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        wb.close()
        return rows
